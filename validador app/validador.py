import customtkinter as ctk
from tkinter import Menu, simpledialog, messagebox, filedialog
import pandas as pd
import json
import os
import pandas as pd
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill



# Ruta del archivo JSON donde se guardarán las áreas
archivo_json = "areas.json"

# Función para cargar las áreas desde el archivo JSON
def cargar_areas():
    if os.path.exists(archivo_json):
        with open(archivo_json, "r") as archivo:
            return json.load(archivo)
    return {}

# Función para guardar las áreas en el archivo JSON
def guardar_areas():
    with open(archivo_json, "w") as archivo:
        json.dump(areas, archivo, indent=4)

# Función para agregar un área
def agregar_area():
    nueva_area = simpledialog.askstring("Agregar Área", "Ingrese el nombre del entorno:")
    if nueva_area:
        if nueva_area in areas:
            messagebox.showerror("Error", "El entorno ya existe.")
            return
        areas[nueva_area] = []  # Cada área comienza con una lista vacía de validadores
        guardar_areas()
        actualizar_botones_areas()

# Función para actualizar los botones de áreas en el panel izquierdo
def actualizar_botones_areas():
    # Limpiar el panel izquierdo
    for widget in frame_izquierdo.winfo_children():
        widget.destroy()

    # Botón para agregar una nueva área
    btn_agregar_area = ctk.CTkButton(
        frame_izquierdo,
        text="+",
        font=ctk.CTkFont(size=20, weight="bold"),
        width=50,
        height=50,
        command=agregar_area
    )
    btn_agregar_area.pack(pady=10)

    # Botones para las áreas existentes
    for area_nombre in areas:
        btn_area = ctk.CTkButton(
            frame_izquierdo,
            text=area_nombre,
            command=lambda nombre=area_nombre: seleccionar_area(nombre)
        )
        btn_area.pack(pady=5, fill="x")

# Función para mostrar opciones del área seleccionada
def seleccionar_area(area):
    # Limpiar el contenido del panel derecho
    for widget in frame_derecho.winfo_children():
        widget.destroy()
    
    # Mostrar las opciones para el área seleccionada
    ctk.CTkLabel(
        frame_derecho,
        text=f"Validadores del Entorno: {area}",
        font=ctk.CTkFont(size=16, weight="bold")
    ).pack(pady=10)
    
    # Mostrar los validadores existentes como botones
    for validador in areas[area]:
        ctk.CTkButton(
            frame_derecho,
            text=validador["nombre"],
            command=lambda v=validador: gestionar_validador(area, v)
        ).pack(pady=5, fill="x")
    
    # Botón para agregar un nuevo validador
    ctk.CTkButton(
        frame_derecho,
        text="Agregar Validador",
        command=lambda: agregar_validador(area)
    ).pack(pady=10)
    
    # Botón para eliminar el área actual
    ctk.CTkButton(
        frame_derecho,
        text="Eliminar Entorno",
        fg_color="red",
        command=lambda: eliminar_area(area)
    ).pack(pady=10)

# Función para agregar un validador a un área
def agregar_validador(area):
    nombre_validador = simpledialog.askstring("Agregar Validador", "Ingrese el nombre del validador:")
    if not nombre_validador:
        return
    
    nuevo_validador = {"nombre": nombre_validador, "reglas": []}  # Inicialmente sin reglas
    areas[area].append(nuevo_validador)
    guardar_areas()
    seleccionar_area(area)

# Función para gestionar las reglas de un validador
def gestionar_validador(area, validador):
    # Limpiar el panel derecho
    for widget in frame_derecho.winfo_children():
        widget.destroy()
    
    # Título del validador
    ctk.CTkLabel(
        frame_derecho,
        text=f"Reglas para el validador: {validador['nombre']}",
        font=ctk.CTkFont(size=16, weight="bold")
    ).pack(pady=10)
    
    # Mostrar las reglas existentes
    for i, regla in enumerate(validador["reglas"]):
        ctk.CTkButton(
            frame_derecho,
            text=f"Regla {i + 1}: {regla}",
            command=lambda r=regla: editar_regla(area, validador, r)
        ).pack(pady=5, fill="x")
    
    # Botón para agregar una nueva regla
    ctk.CTkButton(
        frame_derecho,
        text="Agregar Regla",
        command=lambda: agregar_regla(area, validador)
    ).pack(pady=10)
    
    # Botón para analizar un archivo Excel
    ctk.CTkButton(
        frame_derecho,
        text="Analizar Excel",
        command=lambda: analizar_excel(validador)
    ).pack(pady=10)
    
    # Botón para volver a la lista de validadores
    ctk.CTkButton(
        frame_derecho,
        text="Volver",
        command=lambda: seleccionar_area(area)
    ).pack(pady=10)

# Función para agregar una regla a un validador
def agregar_regla(area, validador):
    
   
    
    # Crear una ventana modal
    modal = ctk.CTkToplevel()
    modal.title("Seleccionar Tipo de Regla")
    modal.geometry("300x200")
    modal.grab_set()  # Bloquea la ventana principal hasta que se cierre esta

    tipo_regla_var = ctk.StringVar(value="longitud")  # Valor predeterminado

    ctk.CTkLabel(modal, text="Seleccione el tipo de regla:").pack(pady=10)
    tipo_regla_menu = ctk.CTkOptionMenu(
        modal, 
        values=["longitud", "numerico", "regex", "unico", "dependiente_positivo", "dependiente_error" ,"no_vacio", "dependiente longitud", "dependiente edad"], 
        variable=tipo_regla_var
    )
    tipo_regla_menu.pack(pady=10)

    def confirmar_tipo_regla():
        tipo_regla = tipo_regla_var.get()
        modal.destroy()  # Cerrar la ventana modal
        
        if tipo_regla == "longitud":
            
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar por longitud (por ejemplo, Cedula):")
            if not columna:
                return
            condicion = simpledialog.askstring("Longitud", "Ingrese la longitud máxima (ejemplo: 10):")
            
            if not condicion:
                return
            nueva_regla = {"columna": columna, "tipo": "longitud", "condicion": f"<= {condicion}"}
        
        elif tipo_regla == "numerico":
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar numerico (por ejemplo, Telefono):")
            if not columna:
                return
            condicion = simpledialog.askstring("Numerico", "Ingrese la condición (ejemplo: 'mayor  5'):")
            if not condicion:
                return
            nueva_regla = {"columna": columna, "tipo": "numerico", "condicion": condicion}
        
        elif tipo_regla == "regex":
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar para que no tenga caracteres especiales (por ejemplo, Nombres):")
            if not columna:
                return
            patron = simpledialog.askstring("Expresión Regular", "Ingrese el patrón regex (ejemplo: \\d{3}-\\d{2}-\\d{4}):")
            if not patron:
                return
            nueva_regla = {"columna": columna, "tipo": "regex", "patron": patron}
        
        elif tipo_regla == "unico":
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar para qvalores unicos (por ejemplo, Nombres):")
            if not columna:
                return  
            nueva_regla = {"columna": columna, "tipo": "unico"}
        
        elif tipo_regla == "dependiente_positivo":
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar  (por ejemplo, Telefono):")
            if not columna:
                return
            
            columna_dependiente = simpledialog.askstring("Columna Dependiente", "¿De qué columna depende esta regla? (por ejemplo, A):")
            if not columna_dependiente:
                return
            valor_dependiente = simpledialog.askstring("Valor Dependiente", "¿Qué valor debe tener la columna dependiente? (ejemplo: 50):")
            if not valor_dependiente:
                return
            valor_dependiente = float(valor_dependiente) if valor_dependiente.replace('.', '', 1).isdigit() else valor_dependiente
            
            valor_esperado = simpledialog.askstring("Valor Esperado", "¿Qué valor debe tener la columna a validar si la columna dependiente tiene este valor? (ejemplo: 51):")
            
            if not valor_esperado:
                return
            
            nueva_regla = {
                "columna": columna, 
                "tipo": "dependiente positivo", 
                "columna_dependiente": columna_dependiente, 
                "valor_dependiente": valor_dependiente, 
                "valor_esperado": valor_esperado
            }
            
        elif tipo_regla == "dependiente_error":
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar numerico (por ejemplo, Telefono):")
            if not columna:
                return
            
            columna_dependiente = simpledialog.askstring("Columna Dependiente", "¿De qué columna depende esta regla? (por ejemplo, A):")
            if not columna_dependiente:
                return
            valor_dependiente = simpledialog.askstring("Valor Dependiente", "¿Qué valor debe tener la columna dependiente? (ejemplo: VEN):")
            if not valor_dependiente:
                return
            valor_dependiente = float(valor_dependiente) if valor_dependiente.replace('.', '', 1).isdigit() else valor_dependiente
            
            valor_esperado = simpledialog.askstring("Valor Esperado", "¿Qué valor debe tener la columna a validar si la columna dependiente tiene este valor? (ejemplo: NO APLICA):")
            if not valor_esperado:
                return
            
            nueva_regla = {
                "columna": columna, 
                "tipo": "dependiente_error", 
                "columna_dependiente": columna_dependiente, 
                "valor_dependiente": valor_dependiente, 
                "valor_esperado": valor_esperado
            }
            
            
        elif tipo_regla == "no_vacio":
            columnas = simpledialog.askstring(
                "No Vacío", 
                "Ingrese las columnas que no pueden estar vacías, separadas por comas (ejemplo: A, B, C):"
            )
            if not columnas:
                return
            columna = "Ficha_fic"
            columnas = [col.strip() for col in columnas.split(",") if col.strip()]
            nueva_regla = {"columna": columna, "tipo": "no_vacio", "columnas": columnas}
        
        
        elif tipo_regla == "dependiente longitud":
            
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar (por ejemplo, DOCUMENTO):")
            if not columna:
                return
            
            columna_dependiente = simpledialog.askstring("Columna Dependiente", "¿De qué columna depende esta regla? (por ejemplo, TIPO DOCUMENTO ):")
            if not columna_dependiente:
                return 
            
            valor_dependiente = simpledialog.askstring("Valor Dependiente", "¿Qué valor debe tener la columna dependiente? (ejemplo: 3- TI):")
            if not valor_dependiente:
                return
            
            valor_esperado = simpledialog.askstring("Valor Esperado", "Que cantidad de digitos debe tener la columna a validar (por ejemplo: 10)")
            if not valor_esperado:
                return
            
            nueva_regla = {
                "columna": columna,
                "tipo": "dependiente longitud",
                "columna_dependiente": columna_dependiente,
                "valor_dependiente": valor_dependiente, 
                "valor_esperado": f"<= {valor_esperado}"
            }
        
        elif tipo_regla == "dependiente edad":
            
            columna = simpledialog.askstring("Agregar Regla", "Ingrese la columna a validar (por ejemplo, ESTADO CIVIL):")
            if not columna:
                return
            
            columna_dependiente = simpledialog.askstring("Columna Dependiente", "¿De qué columna depende esta regla? (por ejemplo, FECHA DE NACIMIENTO ):")
            if not columna_dependiente:
                return 
            
            valor_dependiente = simpledialog.askstring("Valor Dependiente", "indique la edad o rango de edades separados por coma mierda (por ejemplo: 7,17 )")
            if not valor_dependiente:
                return
            
            valor_esperado = simpledialog.askstring("Valor Esperado", "valor esperado segun la edad")
            if not valor_esperado:
                return
            
            nueva_regla = {
                "columna": columna,
                "tipo": "dependiente longitud",
                "columna_dependiente": columna_dependiente,
                "valor_dependiente": valor_dependiente, 
                "valor_esperado": valor_esperado
            }
                    
        else:
            messagebox.showerror("Error", "Tipo de regla no reconocido.")
            return

        validador["reglas"].append(nueva_regla)
        guardar_areas()
        gestionar_validador(area, validador)

    # Botón para confirmar la selección
    confirmar_btn = ctk.CTkButton(modal, text="Confirmar", command=confirmar_tipo_regla)
    confirmar_btn.pack(pady=20)

    modal.protocol("WM_DELETE_WINDOW", modal.destroy)  # Permite cerrar la ventana con la 'X'


# Función para editar una regla
def editar_regla(area, validador, regla):
    nueva_regla = simpledialog.askstring("Editar Regla", f"Modificar regla: {regla}")
    if nueva_regla:
        indice = validador["reglas"].index(regla)
        validador["reglas"][indice] = nueva_regla
        guardar_areas()
        gestionar_validador(area, validador)


def analizar_excel(validador):
    archivo_excel = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")]
    )
    if archivo_excel:
        try:
            # Leer el archivo Excel
            df = pd.read_excel(archivo_excel)

            # Cargar el archivo Excel en openpyxl para aplicar formato
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active

            # Color de fondo rojo para las celdas que no cumplen con la condición
            rojo_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for regla in validador["reglas"]:
                columna = regla.get("columna")
               
                
                tipo = regla.get("tipo")

                if columna in df.columns:
                    col_idx = df.columns.get_loc(columna) + 1  # Obtener el índice de la columna en openpyxl (1-based)
                    
                    if tipo == "longitud":
                        max_longitud = int(regla["condicion"].split("<= ")[1])
                        violaciones = df[columna][df[columna].astype(str).str.len() > max_longitud]
                        for idx in violaciones.index:
                            # Marcar en rojo las celdas que violan la regla de longitud
                            ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # +2 por el encabezado
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                            

                    elif tipo == "numerico":
                        try:
                            operador, valor = regla["condicion"].split(" ")
                            valor = int(valor)

                            if operador == "mayor":
                                violaciones = df[columna][df[columna] > valor]
                            elif operador == "menor":
                                violaciones = df[columna][df[columna] < valor]

                            for idx in violaciones.index:
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                

                        except ValueError:
                            pass

                    elif tipo == "regex":
                        patron = regla["patron"]
                        # Normalizar los datos
                        df[columna] = df[columna].astype(str).str.strip()
                        
                        # Filtrar las filas que no cumplen con el patrón
                        violaciones = df[columna][df[columna].str.fullmatch(patron) == False]
                        
                        for idx in violaciones.index:
                            ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # Marcar en rojo
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                            

                    elif tipo == "unico":
                        duplicados = df[columna][df[columna].duplicated()]
                        for idx in duplicados.index:
                            ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                            # Marcar en rojo

                    elif tipo == "dependiente positivo":
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        valor_esperado = regla.get("valor_esperado")
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1

                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]

                            # Filtrar las filas que NO cumplen con el valor esperado en la columna principal
                            violaciones = filas_dependientes[filas_dependientes[columna] != valor_esperado]

                            # Solo marcar en rojo las filas que no cumplen con la condición
                            for idx in violaciones.index:
                                # Marcar en rojo las celdas que no cumplen la condición (solo las filas con violaciones)
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
                                
                                
                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                            
                    elif tipo == "no_vacio":
                        columnas = regla.get("columnas")
                        print("Columnas a verificar:", columnas)  # Imprimir para verificar las columnas

                        # Asegúrate de que 'columna' sea una lista
                        if isinstance(columnas, str):  # Si 'columna' es una cadena en lugar de lista
                            columnas = [columnas]  # Convertirla en una lista
                        
                        for columna in columnas:
                            if columna in df.columns:
                                col_idx = df.columns.get_loc(columna) + 1  # Obtener el índice de la columna en openpyxl (1-based)
                                print(f"Índice de columna '{columna}': {col_idx}")
                                # Filtrar las filas que tienen valores vacíos en la columna
                                violaciones = df[df[columna].isnull() | (df[columna] == "")]
                                print(f"Violaciones encontradas en columna '{columna}': {violaciones.index.tolist()}")
                                for idx in violaciones.index:
                                    print(f"Marcando fila {idx} en columna {columna}")  # Imprimir para depurar
                                    # Marcar en rojo las celdas que tienen valores vacíos en la columna principal
                                    ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                    ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # Marcar en rojo otra celda relacionada, si es necesario
                            else:
                                messagebox.showinfo("Advertencia", f"Columna '{columna}' no encontrada en el archivo Excel.")

                    elif tipo == "dependiente_error":
                    
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        valor_esperado = regla.get("valor_esperado")
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1

                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]

                            # Filtrar las filas que NO cumplen con el valor esperado en la columna principal
                            violaciones = filas_dependientes[filas_dependientes[columna] == valor_esperado]

                            # Solo marcar en rojo las filas que no cumplen con la condición
                            for idx in violaciones.index:
                                # Marcar en rojo las celdas que no cumplen la condición (solo las filas con violaciones)
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
                                
                              
                                
                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                        
                    elif tipo == "dependiente longitud":
                        print("hola")
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        valor_esperado = regla.get("valor_esperado")
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1

                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]
                            
                            max_longitud = int(regla["valor_esperado"].split("<= ")[1])
                            
                            violaciones = filas_dependientes[filas_dependientes[columna] .astype(str).str.len() > max_longitud]
                            
                            for idx in violaciones.index:
                                # Marcar en rojo las celdas que violan la regla de longitud
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # +2 por el encabezado
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
      
                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                    
                    elif tipo == "dependiente edad":
                       
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1
                        
                        columna_dependiente = regla.get("columna_dependiente") # estado civil
                        valor_dependiente = regla.get("valor_dependiente") # fecha de nacimiento
                        valor_esperado = regla.get("valor_esperado")
                        
                        # calcular edad 
                        

                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]
                            
                            max_longitud = int(regla["valor_esperado"].split("<= ")[1])
                            
                            violaciones = filas_dependientes[filas_dependientes[columna] .astype(str).str.len() > max_longitud]
                            
                            for idx in violaciones.index:
                                # Marcar en rojo las celdas que violan la regla de longitud
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # +2 por el encabezado
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
                                

                                
                           
                                
                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                    
                else:
                    
                    messagebox.showinfo("Advertencia", f"Columna '{columna}' no encontrada en el archivo Excel.")

            # Guardar el nuevo archivo Excel con las celdas marcadas
            nuevo_archivo = filedialog.asksaveasfilename(
                title="Guardar archivo Excel con validaciones",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")]
            )

            if nuevo_archivo:
                wb.save(nuevo_archivo)
                messagebox.showinfo("Éxito", "Se ha creado un nuevo archivo con las validaciones marcadas en rojo.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo analizar el archivo Excel:\n{e}")




# Función para eliminar un área
def eliminar_area(area):
    if messagebox.askyesno("Confirmar", f"¿Desea eliminar el área '{area}'?"):
        del areas[area]
        guardar_areas()
        actualizar_botones_areas()
        for widget in frame_derecho.winfo_children():
            widget.destroy()  # Limpiar el panel derecho

# Configuración inicial de la ventana principal
ctk.set_appearance_mode("Dark")  # Modo oscuro
ventana = ctk.CTk()
ventana.title("Gestión de Áreas y Validadores")
ventana.geometry("800x600")

# Crear un menú
menu_bar = Menu(ventana)
ventana.config(menu=menu_bar)

# Menú "Archivo"
menu_archivo = Menu(menu_bar, tearoff=0)
menu_archivo.add_command(label="Nuevo")
menu_archivo.add_command(label="Abrir")
menu_archivo.add_command(label="Guardar")
menu_archivo.add_separator()
menu_archivo.add_command(label="Salir", command=ventana.quit)
menu_bar.add_cascade(label="Archivo", menu=menu_archivo)

# Crear dos paneles: izquierdo y derecho
frame_izquierdo = ctk.CTkFrame(ventana, width=200)
frame_izquierdo.pack(side="left", fill="y", padx=10, pady=10)

frame_derecho = ctk.CTkFrame(ventana)
frame_derecho.pack(side="right", fill="both", expand=True, padx=10, pady=10)

# Cargar las áreas y crear los botones en el panel izquierdo
areas = cargar_areas()
actualizar_botones_areas()

ventana.mainloop()
